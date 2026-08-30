from __future__ import annotations

from io import BytesIO
import re

from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import Engine, text


class ScfMapping(BaseModel):
    """One SCF control reference to an external framework control."""

    framework: str
    version: str
    control_code: str


class ScfControl(BaseModel):
    """One SCF control and its selected roadmap mappings."""

    control_code: str
    title: str
    description: str
    mappings: list[ScfMapping]


_TARGETS = {
    "AICPA TSC 2017:2022 (used for SOC 2)": ("SOC 2 TSC", "2017:2022"),
    "ISO 27001 2022": ("ISO 27001", "2022"),
    "NIST 800-53 R5": ("NIST SP 800-53", "5.2.0"),
    "PCI DSS 4.0.1": ("PCI DSS", "4.0.1"),
    "US HIPAA Security Rule / NIST SP 800-66 R2": ("HIPAA Security Rule", "2013"),
}


def _header(value: object) -> str:
    return " ".join(str(value or "").split())


def _nist_code(code: str) -> str:
    return re.sub(r"(?<=-|\.)0+(?=\d)", "", code)


def parse_scf(payload: bytes) -> list[ScfControl]:
    """Parse roadmap-priority crosswalks from an official SCF workbook."""
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    sheet = next(workbook[name] for name in workbook.sheetnames if name.startswith("SCF 20"))
    rows = sheet.iter_rows(values_only=True)
    headers = [_header(value) for value in next(rows)]
    columns = {header: index for index, header in enumerate(headers)}
    targets = [(columns[header], *target) for header, target in _TARGETS.items()]
    controls = []

    for row in rows:
        code = row[columns["SCF #"]]
        if not code:
            continue
        mappings = []
        for index, framework, version in targets:
            for mapped_code in str(row[index] or "").splitlines():
                mapped_code = mapped_code.strip()
                if mapped_code:
                    mappings.append(
                        ScfMapping(
                            framework=framework,
                            version=version,
                            control_code=_nist_code(mapped_code)
                            if framework == "NIST SP 800-53"
                            else mapped_code,
                        )
                    )
        controls.append(
            ScfControl(
                control_code=str(code),
                title=str(row[columns["SCF Control"]]),
                description=str(row[columns["Secure Controls Framework (SCF) Control Description"]]),
                mappings=mappings,
            )
        )
    return controls


def ingest_scf(controls: list[ScfControl], engine: Engine, *, source_url: str) -> int:
    """Upsert SCF controls and selected external-framework mappings."""
    with engine.begin() as connection:
        framework_sql = text(
            "INSERT INTO frameworks (name, version, publisher, machine_readable_source) "
            "VALUES (:name, :version, :publisher, :source) "
            "ON CONFLICT (name, version) DO UPDATE SET machine_readable_source = EXCLUDED.machine_readable_source "
            "RETURNING id"
        )
        control_sql = text(
            "INSERT INTO controls (framework_id, control_code, title, description) "
            "VALUES (:framework_id, :code, :title, :description) "
            "ON CONFLICT (framework_id, control_code) DO UPDATE SET title = EXCLUDED.title, "
            "description = EXCLUDED.description RETURNING id"
        )
        scf_id = connection.execute(
            framework_sql,
            {"name": "Secure Controls Framework", "version": "2026.2", "publisher": "SCF Council", "source": source_url},
        ).scalar_one()
        framework_ids = {
            target: connection.execute(
                framework_sql,
                {"name": target, "version": version, "publisher": target, "source": source_url},
            ).scalar_one()
            for target, version in _TARGETS.values()
        }
        count = 0
        for control in controls:
            source_id = connection.execute(
                control_sql,
                {"framework_id": scf_id, "code": control.control_code, "title": control.title, "description": control.description},
            ).scalar_one()
            for mapping in control.mappings:
                target_id = connection.execute(
                    control_sql,
                    {"framework_id": framework_ids[mapping.framework], "code": mapping.control_code, "title": mapping.control_code, "description": ""},
                ).scalar_one()
                # ponytail: neutral inferred edge; import STRM metadata when SCF exposes it machine-readably.
                connection.execute(
                    text(
                        "INSERT INTO crosswalks (control_a, control_b, relation, strength, source) "
                        "VALUES (:a, :b, 'related', 0.5, 'SCF 2026.2') "
                        "ON CONFLICT (control_a, control_b, source) DO NOTHING"
                    ),
                    {"a": source_id, "b": target_id},
                )
                count += 1
    return count

